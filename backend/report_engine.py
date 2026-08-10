"""
report_engine.py — 4-Sheet Excel & PDF Reporting Layer
=======================================================
UNIBEN Biodiversity Pipeline · Root Module

Thesis title: "Development of a campus-scale biodiversity and environmental
data pipeline using drone imagery and sensor integration for machine learning
application: A case study for UNIBEN Ugbowo campus."

Sheets produced
---------------
  1. Sanitized Telemetry   — cleaned sensor_readings + altitude + rolling MA-5
  2. Micro-Climate Variance — daily summary stats + OWM ambient variance delta
  3. Taxonomic Metadata    — full JOIN field_observations ↔ taxonomic_metadata
  4. Anomaly Flags         — rows where Z-score > 3 sigma (per sensor field)

Usage
-----
    # CLI one-shot:
    python report_engine.py [session_id]

    # Import into FastAPI:
    from report_engine import ReportEngine
    engine = ReportEngine()
    path = engine.generate_excel_spreadsheet(session_id="all")
"""

from __future__ import annotations

import logging
import sqlite3
from datetime import datetime, timezone
from pathlib import Path
from typing import Optional

log = logging.getLogger("report_engine")
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
)

# ---------------------------------------------------------------------------
# Dependency guards
# ---------------------------------------------------------------------------

try:
    import pandas as pd
    _HAS_PANDAS = True
except ImportError:
    _HAS_PANDAS = False

try:
    import numpy as np
    _HAS_NUMPY = True
except ImportError:
    _HAS_NUMPY = False

try:
    import openpyxl                           # noqa: F401
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side,
    )
    from openpyxl.utils import get_column_letter
    _HAS_OPENPYXL = True
except ImportError:
    _HAS_OPENPYXL = False

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

THESIS_TITLE = (
    "Development of a campus-scale biodiversity and environmental data pipeline "
    "using drone imagery and sensor integration for machine learning application: "
    "A case study for UNIBEN Ugbowo campus."
)
DEVELOPER    = "Sanwo Demilade <demisanwo004@gmail.com>"

BASE_DIR    = Path(__file__).resolve().parent
DB_PATH     = BASE_DIR / "data" / "biodiversity.db"
EXPORTS_DIR = BASE_DIR / "data" / "exports"

# ARGB colour palette (openpyxl uses ARGB without #)
CLR_HEADER_FOREST = "FF064E3B"   # deep forest green — primary header
CLR_HEADER_MID    = "FF065F46"   # mid forest — secondary header
CLR_HEADER_ACCENT = "FF34D399"   # emerald accent — sheet-2/3/4 headers
CLR_HEADER_ALERT  = "FF7C3AED"   # violet — anomaly sheet header
CLR_ROW_ALT       = "FFE8FDF5"   # faint mint — alternate data rows
CLR_ROW_PLAIN     = "FFFFFFFF"
CLR_FONT_WHITE    = "FFFFFFFF"
CLR_FONT_DARK     = "FF111827"
CLR_ANOMALY_ROW   = "FFFFF3CD"   # amber tint — anomaly rows

SENSOR_FIELDS = [
    "temperature_c", "humidity_percent", "pressure_hPa",
    "light_lux", "sound_db", "altitude_m",
]

THIN_BORDER = None  # initialised lazily after openpyxl imports succeed


def _make_thin_border():
    s = Side(border_style="thin", color="FFD1FAE5")
    return Border(left=s, right=s, top=s, bottom=s)


# ---------------------------------------------------------------------------
# Style helpers
# ---------------------------------------------------------------------------

def _hdr_cell(ws, row: int, col: int, value: str,
               bg: str = CLR_HEADER_FOREST,
               font_size: int = 9,
               bold: bool = True,
               wrap: bool = True) -> None:
    """Write and style a single header cell."""
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(name="Calibri", bold=bold, color=CLR_FONT_WHITE, size=font_size)
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=wrap)
    cell.border = _make_thin_border()


def _data_cell(ws, row: int, col: int, value, alt_row: bool = False, num_fmt: Optional[str] = None) -> None:
    """Write and style a data cell."""
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(name="Calibri", size=8, color=CLR_FONT_DARK)
    cell.fill = PatternFill("solid", fgColor=CLR_ROW_ALT if alt_row else CLR_ROW_PLAIN)
    cell.alignment = Alignment(horizontal="left", vertical="center")
    cell.border = _make_thin_border()
    if num_fmt:
        cell.number_format = num_fmt


def _title_banner(ws, title: str, subtitle: str, num_cols: int) -> None:
    """Insert a merged two-row title banner at the top of a sheet."""
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_cols)
    cell = ws.cell(row=1, column=1, value=title)
    cell.font      = Font(name="Calibri", bold=True, color=CLR_FONT_WHITE, size=12)
    cell.fill      = PatternFill("solid", fgColor=CLR_HEADER_FOREST)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 40

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=num_cols)
    cell2 = ws.cell(row=2, column=1, value=subtitle)
    cell2.font      = Font(name="Calibri", italic=True, color=CLR_FONT_WHITE, size=8)
    cell2.fill      = PatternFill("solid", fgColor=CLR_HEADER_MID)
    cell2.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 16


def _auto_col_width(ws) -> None:
    """Set comfortable auto column widths (max 45 chars)."""
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            try:
                max_len = max(max_len, len(str(cell.value or "")))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 4, 45)


# ---------------------------------------------------------------------------
# ReportEngine
# ---------------------------------------------------------------------------

class ReportEngine:
    """
    Generates styled multi-sheet Excel (.xlsx) workbooks from the live SQLite
    database.  All sheets use the thesis title banner and the UNIBEN green
    colour palette.
    """

    def __init__(
        self,
        db_path:     Optional[Path] = None,
        exports_dir: Optional[Path] = None,
    ) -> None:
        self.db_path     = Path(db_path)     if db_path     else DB_PATH
        self.exports_dir = Path(exports_dir) if exports_dir else EXPORTS_DIR
        self.exports_dir.mkdir(parents=True, exist_ok=True)

    # ── DB helpers ────────────────────────────────────────────────────────

    def _conn(self) -> sqlite3.Connection:
        c = sqlite3.connect(self.db_path)
        c.row_factory = sqlite3.Row
        return c

    # ── Sheet 1: Sanitized Telemetry ─────────────────────────────────────

    def _build_sheet1(self, wb, session_id: str) -> None:
        """
        Sheet 1 — Sanitized Real-Time & Fallback Telemetry.

        Adds rolling 5-point moving average columns for each sensor field.
        Anomaly source tags: LIVE_ESP32, ESP32_SD_CARD.
        """
        if not _HAS_PANDAS or not _HAS_NUMPY:
            wb.create_sheet("Sanitized Telemetry")
            return

        with self._conn() as conn:
            df = pd.read_sql_query(
                "SELECT id, device_id, temperature_c, humidity_percent, pressure_hPa, "
                "light_lux, sound_db, altitude_m, latitude, longitude, "
                "observed_at, data_source, notes "
                "FROM sensor_readings ORDER BY observed_at ASC",
                conn,
            )

        if df.empty:
            ws = wb.create_sheet("Sanitized Telemetry")
            ws.cell(row=1, column=1, value="No sensor data available.")
            return

        # Rolling 5-point moving average
        for col in SENSOR_FIELDS:
            if col in df.columns:
                df[f"{col}_ma5"] = df[col].rolling(window=5, min_periods=1).mean().round(4)

        ws = wb.create_sheet("Sanitized Telemetry")

        cols = (
            list(df.columns[:df.columns.get_loc("notes") + 1])
            + [f"{c}_ma5" for c in SENSOR_FIELDS if f"{c}_ma5" in df.columns]
        )
        num_cols = len(cols)

        _title_banner(ws,
            title=THESIS_TITLE,
            subtitle=f"Sheet 1: Sanitized Sensor Telemetry — session: {session_id} | {DEVELOPER}",
            num_cols=num_cols)

        for ci, col_name in enumerate(cols, start=1):
            label = col_name.replace("_", " ").title().replace("Hpa", "hPa").replace("Db", "dB").replace("Ma5", "MA-5")
            _hdr_cell(ws, row=3, col=ci, value=label, bg=CLR_HEADER_MID)

        for ri, (_, row) in enumerate(df[cols].iterrows(), start=4):
            alt = (ri - 4) % 2 == 1
            for ci, col_name in enumerate(cols, start=1):
                val = row[col_name]
                fmt = "0.0000" if "ma5" in col_name else None
                if pd.isna(val):
                    val = None
                _data_cell(ws, row=ri, col=ci, value=val, alt_row=alt, num_fmt=fmt)

        ws.freeze_panes = "A4"
        _auto_col_width(ws)

    # ── Sheet 2: Micro-Climate Variance ───────────────────────────────────

    def _build_sheet2(self, wb) -> None:
        """
        Sheet 2 — Micro-Climate Summary Statistics & Variance Analytics.

        Joins sensor_readings daily stats with external_weather_metadata OWM data.
        """
        if not _HAS_PANDAS:
            wb.create_sheet("Micro-Climate Variance")
            return

        ws = wb.create_sheet("Micro-Climate Variance")

        with self._conn() as conn:
            df = pd.read_sql_query(
                """
                SELECT
                    DATE(observed_at)     AS date,
                    COUNT(*)              AS readings_count,
                    ROUND(AVG(temperature_c), 3)    AS temp_avg,
                    ROUND(MIN(temperature_c), 3)    AS temp_min,
                    ROUND(MAX(temperature_c), 3)    AS temp_max,
                    ROUND(STDDEV_APPROX(temperature_c), 3) AS temp_stddev,
                    ROUND(AVG(humidity_percent), 3) AS humidity_avg,
                    ROUND(AVG(pressure_hPa), 3)     AS pressure_avg,
                    ROUND(AVG(light_lux), 3)        AS light_avg,
                    ROUND(AVG(sound_db), 3)         AS sound_avg,
                    ROUND(AVG(altitude_m), 3)       AS altitude_avg
                FROM sensor_readings
                GROUP BY DATE(observed_at)
                ORDER BY date ASC
                """,
                conn,
            )
            owm_df = pd.read_sql_query(
                """
                SELECT
                    DATE(fetch_timestamp) AS date,
                    ROUND(AVG(ambient_temp_c), 3) AS owm_temp_avg,
                    ROUND(AVG(aqi), 2)            AS aqi_avg,
                    sky_condition
                FROM external_weather_metadata
                GROUP BY DATE(fetch_timestamp)
                """,
                conn,
            )

        if not df.empty and not owm_df.empty:
            df = df.merge(owm_df, on="date", how="left")
        elif not owm_df.empty:
            df = owm_df

        # Compute sensor↔OWM temperature variance
        if "temp_avg" in df.columns and "owm_temp_avg" in df.columns:
            df["owm_variance_c"] = (df["owm_temp_avg"] - df["temp_avg"]).round(4)

        # SQLite does not have STDDEV natively — fall back to pandas
        if "temp_stddev" not in df.columns or df["temp_stddev"].isna().all():
            with self._conn() as conn2:
                raw = pd.read_sql_query(
                    "SELECT DATE(observed_at) AS date, temperature_c FROM sensor_readings",
                    conn2,
                )
            if not raw.empty:
                std_map = raw.groupby("date")["temperature_c"].std().round(3).to_dict()
                df["temp_stddev"] = df["date"].map(std_map)

        cols = list(df.columns)
        num_cols = len(cols)
        _title_banner(ws,
            title=THESIS_TITLE,
            subtitle=f"Sheet 2: Daily Micro-Climate Summary & OWM Variance | {DEVELOPER}",
            num_cols=num_cols)

        for ci, col_name in enumerate(cols, start=1):
            label = col_name.replace("_", " ").title().replace("Owm", "OWM").replace("Aqi", "AQI")
            _hdr_cell(ws, row=3, col=ci, value=label, bg=CLR_HEADER_ACCENT)

        for ri, (_, row) in enumerate(df.iterrows(), start=4):
            alt = (ri - 4) % 2 == 1
            for ci, col_name in enumerate(cols, start=1):
                val = row[col_name]
                if pd.isna(val) if hasattr(pd, "isna") else val != val:
                    val = None
                fmt = "0.0000" if isinstance(val, float) else None
                _data_cell(ws, row=ri, col=ci, value=val, alt_row=alt, num_fmt=fmt)

        ws.freeze_panes = "A4"
        _auto_col_width(ws)

    # ── Sheet 3: Taxonomic Metadata ───────────────────────────────────────

    def _build_sheet3(self, wb) -> None:
        """
        Sheet 3 — Extracted Taxonomic Metadata & Confidence Indicators.

        Full JOIN: field_observations ↔ taxonomic_metadata (PlantNet results).
        """
        if not _HAS_PANDAS:
            wb.create_sheet("Taxonomic Metadata")
            return

        ws = wb.create_sheet("Taxonomic Metadata")

        with self._conn() as conn:
            df = pd.read_sql_query(
                """
                SELECT
                    fo.obs_id,
                    fo.date,
                    fo.time,
                    fo.campus_zone,
                    fo.common_name,
                    fo.data_source,
                    fo.annotation_confidence AS cv_confidence,
                    fo.observer_id,
                    fo.gps_lat,
                    fo.gps_long,
                    COALESCE(tm.source_api, 'cv_inference')  AS taxonomy_source,
                    COALESCE(tm.kingdom, fo.kingdom)         AS kingdom,
                    COALESCE(tm.phylum,  fo.phylum)          AS phylum,
                    COALESCE(tm.class,   fo.class)           AS class,
                    COALESCE(tm.order_name, fo.order_name)   AS order_name,
                    COALESCE(tm.family,  fo.family)          AS family,
                    COALESCE(tm.genus,   fo.genus)           AS genus,
                    COALESCE(tm.species, fo.species)         AS species,
                    tm.confidence                            AS plantnet_confidence
                FROM field_observations fo
                LEFT JOIN taxonomic_metadata tm ON tm.obs_id = fo.obs_id
                ORDER BY fo.obs_id DESC
                """,
                conn,
            )

        cols = list(df.columns) if not df.empty else [
            "obs_id", "date", "kingdom", "phylum", "class",
            "order_name", "family", "genus", "species", "confidence",
        ]
        num_cols = len(cols)

        _title_banner(ws,
            title=THESIS_TITLE,
            subtitle=f"Sheet 3: Taxonomic Classification Inventory (PlantNet + CV Inference) | {DEVELOPER}",
            num_cols=num_cols)

        for ci, col_name in enumerate(cols, start=1):
            label = col_name.replace("_", " ").title()
            _hdr_cell(ws, row=3, col=ci, value=label, bg=CLR_HEADER_ACCENT, bg2=CLR_HEADER_MID)

        if df.empty:
            ws.cell(row=4, column=1, value="No taxonomic observations recorded yet.")
            _auto_col_width(ws)
            return

        for ri, (_, row) in enumerate(df.iterrows(), start=4):
            alt = (ri - 4) % 2 == 1
            for ci, col_name in enumerate(cols, start=1):
                val = row[col_name]
                if hasattr(pd, "isna") and pd.isna(val):
                    val = None
                fmt = "0.00%" if "confidence" in col_name else None
                _data_cell(ws, row=ri, col=ci, value=val, alt_row=alt, num_fmt=fmt)

        ws.freeze_panes = "A4"
        _auto_col_width(ws)

    # ── Sheet 4: Anomaly Flags ────────────────────────────────────────────

    def _build_sheet4(self, wb) -> None:
        """
        Sheet 4 — Isolated Sensor Anomalies and Outlier Flags.

        Applies Z-score thresholding (|z| > 3) per sensor field.
        Rows are highlighted in amber.
        """
        if not _HAS_PANDAS or not _HAS_NUMPY:
            wb.create_sheet("Anomaly Flags")
            return

        ws = wb.create_sheet("Anomaly Flags")

        with self._conn() as conn:
            df = pd.read_sql_query(
                "SELECT id, device_id, temperature_c, humidity_percent, "
                "pressure_hPa, light_lux, sound_db, altitude_m, "
                "observed_at, data_source "
                "FROM sensor_readings ORDER BY observed_at ASC",
                conn,
            )

        if df.empty:
            ws.cell(row=1, column=1, value="No sensor data available for anomaly analysis.")
            _auto_col_width(ws)
            return

        anomaly_mask = pd.Series([False] * len(df))
        z_scores: dict = {}

        for field in SENSOR_FIELDS:
            if field not in df.columns:
                continue
            series = df[field].dropna()
            if len(series) < 4:
                continue
            mean, std = series.mean(), series.std()
            if std == 0:
                continue
            z = ((df[field] - mean) / std).abs()
            z_scores[field] = z.round(4)
            anomaly_mask |= (z > 3)

        anomaly_df = df[anomaly_mask].copy()
        for field, z_col in z_scores.items():
            col_name = f"z_{field}"
            anomaly_df[col_name] = z_col[anomaly_mask].values

        cols = list(anomaly_df.columns)
        num_cols = max(len(cols), 6)

        _title_banner(ws,
            title=THESIS_TITLE,
            subtitle=f"Sheet 4: Sensor Anomaly Flags — Z-score |z| > 3σ threshold | {DEVELOPER}",
            num_cols=num_cols)

        for ci, col_name in enumerate(cols, start=1):
            label = col_name.replace("_", " ").title().replace("Hpa", "hPa").replace("Db", "dB")
            _hdr_cell(ws, row=3, col=ci, value=label, bg=CLR_HEADER_ALERT)

        if anomaly_df.empty:
            ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=num_cols)
            good_cell = ws.cell(row=4, column=1, value="✅ No anomalies detected. All readings within ±3σ.")
            good_cell.font = Font(name="Calibri", bold=True, size=10, color="FF064E3B")
            good_cell.alignment = Alignment(horizontal="center")
        else:
            for ri, (_, row) in enumerate(anomaly_df.iterrows(), start=4):
                for ci, col_name in enumerate(cols, start=1):
                    val = row[col_name]
                    if hasattr(pd, "isna") and pd.isna(val):
                        val = None
                    cell = ws.cell(row=ri, column=ci, value=val)
                    cell.font  = Font(name="Calibri", size=8, color=CLR_FONT_DARK)
                    cell.fill  = PatternFill("solid", fgColor=CLR_ANOMALY_ROW)
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                    cell.border = _make_thin_border()

        ws.freeze_panes = "A4"
        _auto_col_width(ws)

    # ── Main entry point ──────────────────────────────────────────────────

    def generate_excel_spreadsheet(self, session_id: str = "all") -> Path:
        """
        Build and save a fully styled 4-sheet Excel workbook.

        Returns the absolute Path to the saved .xlsx file.
        """
        if not _HAS_OPENPYXL:
            raise ImportError(
                "openpyxl is required. Install with: pip install openpyxl"
            )
        if not _HAS_PANDAS:
            raise ImportError(
                "pandas is required. Install with: pip install pandas"
            )

        import openpyxl
        wb = openpyxl.Workbook()

        # Remove default 'Sheet' created by openpyxl
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

        log.info("Building Sheet 1 — Sanitized Telemetry …")
        self._build_sheet1(wb, session_id)

        log.info("Building Sheet 2 — Micro-Climate Variance …")
        self._build_sheet2(wb)

        log.info("Building Sheet 3 — Taxonomic Metadata …")
        self._build_sheet3(wb)

        log.info("Building Sheet 4 — Anomaly Flags …")
        self._build_sheet4(wb)

        ts       = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
        filename = f"UNIBEN_BiodiversityReport_{session_id}_{ts}.xlsx"
        out_path = self.exports_dir / filename
        wb.save(str(out_path))
        log.info("Excel workbook saved → %s", out_path)
        return out_path

    def generate_academic_pdf(self) -> Optional[Path]:
        """
        Generate an academic PDF summary using ReportLab.
        Returns None gracefully if ReportLab is not installed.
        """
        try:
            from reportlab.lib.pagesizes import A4
            from reportlab.lib.styles import getSampleStyleSheet
            from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer
            from reportlab.lib.units import cm
        except ImportError:
            log.warning("reportlab not installed — PDF export skipped.")
            return None

        ts       = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
        filename = f"UNIBEN_ThesisReport_{ts}.pdf"
        out_path = self.exports_dir / filename

        doc    = SimpleDocTemplate(str(out_path), pagesize=A4)
        styles = getSampleStyleSheet()
        story  = []

        story.append(Paragraph(THESIS_TITLE, styles["Title"]))
        story.append(Spacer(1, 0.5 * cm))
        story.append(Paragraph(f"Developer: {DEVELOPER}", styles["Normal"]))
        story.append(Spacer(1, 0.5 * cm))
        story.append(Paragraph(f"Generated: {datetime.now(timezone.utc).isoformat()}", styles["Normal"]))
        story.append(Spacer(1, 1 * cm))

        # Pull key counts from DB
        with self._conn() as conn:
            sr_count = conn.execute("SELECT COUNT(*) FROM sensor_readings").fetchone()[0]
            obs_count = conn.execute("SELECT COUNT(*) FROM field_observations").fetchone()[0]
            tax_count = conn.execute("SELECT COUNT(*) FROM taxonomic_metadata").fetchone()[0]

        story.append(Paragraph("Dataset Summary", styles["Heading2"]))
        story.append(Paragraph(f"• Sensor readings: {sr_count}", styles["Normal"]))
        story.append(Paragraph(f"• Field observations: {obs_count}", styles["Normal"]))
        story.append(Paragraph(f"• PlantNet taxonomy records: {tax_count}", styles["Normal"]))

        doc.build(story)
        log.info("Academic PDF saved → %s", out_path)
        return out_path


# ---------------------------------------------------------------------------
# CLI entry point
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    import sys
    session = sys.argv[1] if len(sys.argv) > 1 else "all"
    engine  = ReportEngine()
    path    = engine.generate_excel_spreadsheet(session_id=session)
    print(f"Report generated: {path}")
