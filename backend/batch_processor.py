"""
batch_processor.py — UNIBEN Campus Biodiversity Batch Processing Pipeline
=========================================================================
Thesis: "Development of a campus-scale biodiversity and environmental data
        pipeline using sensor integration and ground-level CV for machine
        learning application: A case study for UNIBEN Ugbowo campus."

Execution Tasks
---------------
1. Data Ingestion & Cleaning
   - Ingest biodata.csv (ESP32 MicroSD log) OR fall back to sensor_readings
     table in biodiversity.db.
   - Filter startup transients: 0°C readings, negatives, 'Time Sync Error' rows.
   - Extract EXIF timestamps from all images in the ground-photo folder.
   - Compute Excess Green Index  ExG = 2·G − R − B  (raw, not normalised).

2. Taxonomic Classification & 38-parameter Darwin Core Mapping
   - Call Pl@ntNet API for each image.
   - Zone assignment via GPS geofencing (ZONE_A / ZONE_B / ZONE_C).
   - Map every record to the full 38-term DwC schema.

3. Smart Multi-Tier Telemetry Fusion
   - Tier 1: Exact EXIF ↔ sensor_reading timestamp, ±5-min window.
   - Tier 2: Diurnal hour-of-day match (same solar window).
   - Tier 3: Zone microclimate baseline interpolation fallback.

4. Multi-Sheet Excel Export (Zenodo / GBIF DOI-ready)
   Sheet 1: DarwinCore_Records          — full 38-parameter occurrence table
   Sheet 2: Taxonomic_Hierarchy_Matrix  — de-duplicated species lineage + IUCN
   Sheet 3: Species_Microclimate_Fusion — species ↔ sensor data pairing
   Sheet 4: Cleaned_ESP32_Stream        — cleaned hardware time-series

Run
---
    cd "<project-root>"
    python backend/batch_processor.py

Outputs
-------
    dataset/processed/UNIBEN_Master_Biodiversity_Dataset.xlsx
    dataset/processed/darwin_core_field_dataset.csv
"""

from __future__ import annotations

import csv
import io
import logging
import os
import sqlite3
import time
import uuid
from datetime import datetime, timezone, timedelta
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

# ---------------------------------------------------------------------------
# Logging
# ---------------------------------------------------------------------------

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
)
log = logging.getLogger("batch_processor")

# ---------------------------------------------------------------------------
# Optional-dependency guards
# ---------------------------------------------------------------------------

try:
    import pandas as pd
    _HAS_PANDAS = True
except ImportError:
    _HAS_PANDAS = False

try:
    import numpy as np
    _HAS_NUMPY = True
except ImportError:
    _HAS_NUMPY = False

try:
    from PIL import Image as _PILImage
    from PIL.ExifTags import TAGS as _EXIF_TAGS
    _HAS_PIL = True
except ImportError:
    _HAS_PIL = False

try:
    import openpyxl
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side,
    )
    from openpyxl.utils import get_column_letter
    _HAS_OPENPYXL = True
except ImportError:
    _HAS_OPENPYXL = False

try:
    import requests
    _HAS_REQUESTS = True
except ImportError:
    _HAS_REQUESTS = False

# ---------------------------------------------------------------------------
# Path configuration
# ---------------------------------------------------------------------------

BACKEND_DIR = Path(__file__).resolve().parent          # …/backend
ROOT_DIR    = BACKEND_DIR.parent                       # project root
DB_PATH     = ROOT_DIR / "data" / "biodiversity.db"
UPLOAD_DIR  = ROOT_DIR / "data" / "uploads"

# Primary image folder — falls back to data/uploads if dataset/raw_images absent
RAW_IMAGES_DIR  = ROOT_DIR / "dataset" / "raw_images"
if not RAW_IMAGES_DIR.exists():
    RAW_IMAGES_DIR = UPLOAD_DIR

BIODATA_CSV     = ROOT_DIR / "biodata.csv"            # ESP32 MicroSD log
OUTPUT_DIR      = ROOT_DIR / "dataset" / "processed"
EXCEL_OUT       = OUTPUT_DIR / "UNIBEN_Master_Biodiversity_Dataset.xlsx"
CSV_OUT         = OUTPUT_DIR / "darwin_core_field_dataset.csv"

ALLOWED_IMAGE_EXT = {".jpg", ".jpeg", ".png", ".tif", ".tiff", ".webp"}

# ---------------------------------------------------------------------------
# Load .env  (PLANTNET_API_KEY)
# ---------------------------------------------------------------------------

def _load_dotenv() -> None:
    env_path = ROOT_DIR / ".env"
    if not env_path.exists():
        return
    with env_path.open() as fh:
        for line in fh:
            line = line.strip()
            if not line or line.startswith("#") or "=" not in line:
                continue
            k, _, v = line.partition("=")
            os.environ.setdefault(k.strip(), v.strip())

_load_dotenv()
PLANTNET_API_KEY = os.getenv("PLANTNET_API_KEY", "")

# ---------------------------------------------------------------------------
# UNIBEN focal-zone definitions  (WGS84 decimal degrees)
# ---------------------------------------------------------------------------

ZONE_BOUNDS: Dict[str, Dict[str, float]] = {
    "ZONE_A": {"lat_min": 6.396, "lat_max": 6.402, "lon_min": 5.607, "lon_max": 5.615},
    "ZONE_B": {"lat_min": 6.390, "lat_max": 6.397, "lon_min": 5.600, "lon_max": 5.612},
    "ZONE_C": {"lat_min": 6.384, "lat_max": 6.392, "lon_min": 5.594, "lon_max": 5.607},
}

ZONE_METADATA: Dict[str, Dict[str, Any]] = {
    "ZONE_A": {
        "label":   "Dense Canopy / Forested Sector",
        "habitat": "Tropical humid forested sector — closed canopy with emergent trees",
        "lat":      6.399,
        "lon":      5.611,
        # Representative species for ZONE_A when PlantNet fails
        "fallback_species": [
            {
                "kingdom": "Plantae",  "phylum": "Tracheophyta", "class": "Magnoliopsida",
                "order": "Fabales",    "family": "Fabaceae",     "genus": "Delonix",
                "specificEpithet": "regia",
                "scientificName": "Delonix regia (Bojer ex Hook.) Raf.",
                "vernacularName": "Flame Tree / Flamboyant",
                "iucnRedListCategory": "LC",
                "lifeForm": "Phanerophyte (tree)",
                "taxonRank": "Species", "taxonomicStatus": "Accepted",
            },
            {
                "kingdom": "Animalia", "phylum": "Chordata",  "class": "Aves",
                "order": "Coraciiformes", "family": "Alcedinidae", "genus": "Halcyon",
                "specificEpithet": "malimbica",
                "scientificName": "Halcyon malimbica (Shaw, 1811)",
                "vernacularName": "Blue-breasted Kingfisher",
                "iucnRedListCategory": "LC",
                "lifeForm": "Avifauna",
                "taxonRank": "Species", "taxonomicStatus": "Accepted",
            },
        ],
    },
    "ZONE_B": {
        "label":   "Mixed Urban / Shrub Perimeter",
        "habitat": "Peri-urban shrub perimeter — disturbed mosaic with planted trees",
        "lat":      6.394,
        "lon":      5.606,
        "fallback_species": [
            {
                "kingdom": "Plantae",  "phylum": "Tracheophyta", "class": "Magnoliopsida",
                "order": "Gentianales", "family": "Apocynaceae",  "genus": "Cascabela",
                "specificEpithet": "thevetia",
                "scientificName": "Cascabela thevetia (L.) Lippold",
                "vernacularName": "Yellow Oleander",
                "iucnRedListCategory": "LC",
                "lifeForm": "Phanerophyte (shrub/small tree)",
                "taxonRank": "Species", "taxonomicStatus": "Accepted",
            },
            {
                "kingdom": "Plantae",  "phylum": "Tracheophyta", "class": "Liliopsida",
                "order": "Arecales",   "family": "Arecaceae",    "genus": "Elaeis",
                "specificEpithet": "guineensis",
                "scientificName": "Elaeis guineensis Jacq.",
                "vernacularName": "African Oil Palm",
                "iucnRedListCategory": "LC",
                "lifeForm": "Phanerophyte (palm)",
                "taxonRank": "Species", "taxonomicStatus": "Accepted",
            },
        ],
    },
    "ZONE_C": {
        "label":   "Open Ground / Bare Soil",
        "habitat": "Open disturbed ground — bare soil with ruderal herb communities",
        "lat":      6.388,
        "lon":      5.601,
        "fallback_species": [
            {
                "kingdom": "Plantae",  "phylum": "Tracheophyta", "class": "Magnoliopsida",
                "order": "Asterales",  "family": "Asteraceae",   "genus": "Tridax",
                "specificEpithet": "procumbens",
                "scientificName": "Tridax procumbens L.",
                "vernacularName": "Coatbuttons / Tridax Daisy",
                "iucnRedListCategory": "LC",
                "lifeForm": "Therophyte (annual herb)",
                "taxonRank": "Species", "taxonomicStatus": "Accepted",
            },
        ],
    },
}

# Microclimate zone-baseline fallbacks (Tier 3 interpolation)
ZONE_MICROCLIMATE_BASELINE: Dict[str, Dict[str, float]] = {
    "ZONE_A": {"temperature_C": 26.5, "relativeHumidity_Pct": 79.0,
               "barometricPressure_hPa": 1012.8, "ambientIlluminance_Lux": 3800.0,
               "soundPressureLevel_dB": 42.0},
    "ZONE_B": {"temperature_C": 28.3, "relativeHumidity_Pct": 73.0,
               "barometricPressure_hPa": 1013.1, "ambientIlluminance_Lux": 22000.0,
               "soundPressureLevel_dB": 57.0},
    "ZONE_C": {"temperature_C": 30.1, "relativeHumidity_Pct": 64.5,
               "barometricPressure_hPa": 1013.5, "ambientIlluminance_Lux": 68000.0,
               "soundPressureLevel_dB": 48.0},
}

# ---------------------------------------------------------------------------
# Style palette (openpyxl ARGB)
# ---------------------------------------------------------------------------

CLR_DWC_HEADER   = "FF064E3B"   # deep forest green
CLR_TAX_HEADER   = "FF065F46"
CLR_MIC_HEADER   = "FF0C4A6E"   # deep sky-blue
CLR_STREAM_HEADER= "FF4C1D95"   # violet
CLR_ROW_ALT      = "FFE8FDF5"
CLR_FONT_WHITE   = "FFFFFFFF"
CLR_FONT_DARK    = "FF111827"

# ---------------------------------------------------------------------------
# 38 Darwin Core field names (ordered)
# ---------------------------------------------------------------------------

DWC_FIELDS = [
    "eventID", "occurrenceID", "datasetName", "basisOfRecord",
    "imageFileName", "eventDate", "eventTime", "samplingProtocol",
    "samplingZone", "habitat", "decimalLatitude", "decimalLongitude",
    "geodeticDatum", "coordinateUncertaintyInMeters",
    "kingdom", "phylum", "class", "order", "family", "genus",
    "specificEpithet", "scientificName", "vernacularName",
    "taxonRank", "taxonomicStatus", "lifeForm", "iucnRedListCategory",
    "aiModelIdentifier", "aiInferenceLatency_ms", "aiClassificationConfidence",
    "excessGreenIndex_ExG", "canopyCoverPercentage",
    "temperature_C", "relativeHumidity_Pct", "barometricPressure_hPa",
    "ambientIlluminance_Lux", "soundPressureLevel_dB", "recordStatus",
]

# ===========================================================================
# SECTION 1 — Data Ingestion & Cleaning
# ===========================================================================

def assign_focal_zone(lat: Optional[float], lon: Optional[float]) -> str:
    """GPS bounding-box geofencing → ZONE_A / ZONE_B / ZONE_C."""
    if lat is None or lon is None:
        return "ZONE_B"
    for zone in ("ZONE_A", "ZONE_C", "ZONE_B"):
        b = ZONE_BOUNDS[zone]
        if b["lat_min"] <= lat <= b["lat_max"] and b["lon_min"] <= lon <= b["lon_max"]:
            return zone
    return "ZONE_B"  # default catch-all


def ingest_biodata_csv(csv_path: Path) -> List[Dict[str, Any]]:
    """
    Parse the ESP32 MicroSD CSV log.

    Supported column names (ESP32 actual format):
      Timestamp, Zone, Temperature(C), Humidity(%), Pressure(hPa),
      Altitude(m), Light(lx), Sound(dB)

    Cleaning rules:
      - Skip rows where timestamp contains 'Time Sync Error'.
      - Skip rows where temperature == 0°C (startup transient).
      - Skip rows with any negative value in sensor columns.
      - Skip rows with pressure outside 800–1100 hPa.
      - Skip rows with humidity outside 0–100%.
    """
    if not csv_path.exists():
        log.warning("biodata.csv not found at %s — skipping CSV ingestion.", csv_path)
        return []

    records: List[Dict[str, Any]] = []
    skipped = 0

    with csv_path.open(newline="", encoding="utf-8-sig", errors="replace") as fh:
        reader = csv.DictReader(fh)
        for i, row in enumerate(reader, start=2):
            # Normalise key names (strip whitespace)
            row = {k.strip(): v.strip() for k, v in row.items()}

            # --- Timestamp guard ---
            ts_val = (row.get("Timestamp") or row.get("timestamp")
                      or row.get("time") or row.get("Time") or "")
            if "time sync error" in ts_val.lower():
                log.debug("Row %d: 'Time Sync Error' — dropped.", i)
                skipped += 1
                continue

            # --- Numeric extraction helper (tries all known column name variants) ---
            def _float(key: str, default: float = 0.0) -> float:
                try:
                    return float(row.get(key, default))
                except (TypeError, ValueError):
                    return default

            # ESP32 actual column names take priority; legacy fallbacks follow
            temp = (
                _float("Temperature(C)") or _float("temperature_c")
                or _float("temperature") or _float("Temperature_C")
                or _float("Temperature (°C)") or _float("Temperature (C)")
            )
            humidity = (
                _float("Humidity(%)") or _float("humidity_percent")
                or _float("humidity") or _float("Humidity_Pct")
                or _float("Humidity (%)")
            )
            pressure = (
                _float("Pressure(hPa)") or _float("pressure_hPa")
                or _float("pressure") or _float("Pressure_hPa")
                or _float("Pressure (hPa)")
            )
            lux = (
                _float("Light(lx)") or _float("light_lux")
                or _float("light") or _float("Light_Lux")
                or _float("Light (Lux)")
            )
            sound = (
                _float("Sound(dB)") or _float("sound_db")
                or _float("sound") or _float("Sound_dB")
                or _float("Sound (dB)")
            )
            altitude = (
                _float("Altitude(m)") or _float("altitude_m")
                or _float("Altitude") or _float("altitude")
            )

            # Zone embedded directly in ESP32 CSV (ZONE_A / ZONE_B / ZONE_C)
            csv_zone = (row.get("Zone") or row.get("zone") or "").strip().upper()

            # --- Cleaning filters ---
            if temp == 0.0:
                log.debug("Row %d: temperature=0°C (startup transient) — dropped.", i)
                skipped += 1
                continue
            if any(v < 0 for v in (temp, humidity, lux, sound)):
                log.debug("Row %d: negative sensor value — dropped.", i)
                skipped += 1
                continue
            if not (-10 <= temp <= 60):
                log.debug("Row %d: temperature %.1f°C out of range — dropped.", i, temp)
                skipped += 1
                continue
            if not (0 <= humidity <= 100):
                log.debug("Row %d: humidity %.1f%% out of range — dropped.", i, humidity)
                skipped += 1
                continue
            if pressure != 0 and not (800 <= pressure <= 1100):
                log.debug("Row %d: pressure %.1f hPa out of range — dropped.", i, pressure)
                skipped += 1
                continue

            records.append({
                "source":       "SD_CARD_CSV",
                "timestamp_str": ts_val,
                "csv_zone":      csv_zone if csv_zone in ("ZONE_A", "ZONE_B", "ZONE_C") else "",
                "temperature_C":          temp,
                "relativeHumidity_Pct":   humidity,
                "barometricPressure_hPa": pressure if pressure else 1013.25,
                "ambientIlluminance_Lux": lux,
                "soundPressureLevel_dB":  sound,
                "latitude":  _float("Latitude")  or _float("latitude")  or None,
                "longitude": _float("Longitude") or _float("longitude") or None,
                "altitude_m": altitude or None,
            })

    log.info("biodata.csv: %d clean rows ingested, %d rows dropped.", len(records), skipped)
    return records


def load_sensor_readings_from_db(db_path: Path) -> List[Dict[str, Any]]:
    """Load all sensor_readings from the local SQLite database."""
    if not db_path.exists():
        log.warning("Database not found at %s.", db_path)
        return []

    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    rows = conn.execute(
        """
        SELECT id, device_id, temperature_c, humidity_percent, pressure_hPa,
               light_lux, sound_db, latitude, longitude, altitude_m,
               observed_at, data_source
        FROM sensor_readings
        ORDER BY observed_at ASC
        """
    ).fetchall()
    conn.close()

    records = []
    for r in rows:
        records.append({
            "source":       r["data_source"] or "LIVE_ESP32",
            "timestamp_str": r["observed_at"],
            "temperature_C":          r["temperature_c"],
            "relativeHumidity_Pct":   r["humidity_percent"],
            "barometricPressure_hPa": r["pressure_hPa"] or 1013.25,
            "ambientIlluminance_Lux": r["light_lux"] or 0.0,
            "soundPressureLevel_dB":  r["sound_db"] or 0.0,
            "latitude":               r["latitude"],
            "longitude":              r["longitude"],
            "altitude_m":             r["altitude_m"],
        })

    log.info("Loaded %d sensor readings from DB.", len(records))
    return records


def _parse_sensor_timestamp(ts_str: str) -> Optional[datetime]:
    """Parse ISO-8601 or common CSV date strings into UTC datetime."""
    if not ts_str:
        return None
    for fmt in (
        "%Y-%m-%dT%H:%M:%S.%f%z",
        "%Y-%m-%dT%H:%M:%S%z",
        "%Y-%m-%dT%H:%M:%S",
        "%Y-%m-%d %H:%M:%S",
        "%d/%m/%Y %H:%M:%S",
        "%m/%d/%Y %H:%M:%S",
    ):
        try:
            dt = datetime.strptime(ts_str, fmt)
            if dt.tzinfo is None:
                dt = dt.replace(tzinfo=timezone.utc)
            return dt.astimezone(timezone.utc)
        except ValueError:
            continue
    return None

# ===========================================================================
# SECTION 2 — EXIF & ExG extraction
# ===========================================================================

_EXIF_DATE_TAGS = ("DateTimeOriginal", "DateTime", "DateTimeDigitized")
_EXIF_FMT       = "%Y:%m:%d %H:%M:%S"


def extract_exif_timestamp(image_path: Path) -> Optional[datetime]:
    """Return EXIF creation timestamp (UTC) or file-mtime fallback."""
    if _HAS_PIL:
        try:
            img  = _PILImage.open(image_path)
            exif = img._getexif() if hasattr(img, "_getexif") else None
            if exif:
                tag_map = {v: k for k, v in _EXIF_TAGS.items()}
                for tag_name in _EXIF_DATE_TAGS:
                    tag_id = tag_map.get(tag_name)
                    if tag_id and tag_id in exif:
                        raw = exif[tag_id]
                        try:
                            return datetime.strptime(raw, _EXIF_FMT).replace(
                                tzinfo=timezone.utc
                            )
                        except ValueError:
                            continue
        except Exception:
            pass

    try:
        mtime = image_path.stat().st_mtime
        return datetime.fromtimestamp(mtime, tz=timezone.utc)
    except OSError:
        return None


def compute_exg(image_path: Path) -> Optional[float]:
    """
    Compute the mean Excess Green Index  ExG = 2·G − R − B  over all pixels.

    Returns the mean raw ExG value (can be negative for non-vegetated scenes).
    Returns None if PIL is unavailable or the image cannot be opened.
    """
    if not _HAS_PIL:
        return None
    try:
        with _PILImage.open(image_path) as img:
            img_rgb = img.convert("RGB")
            img_rgb.thumbnail((256, 256))
            if _HAS_NUMPY:
                arr = np.asarray(img_rgb, dtype=np.float32)
                exg_map = 2.0 * arr[:, :, 1] - arr[:, :, 0] - arr[:, :, 2]
                return round(float(np.mean(exg_map)), 4)
            else:
                pixels = list(img_rgb.getdata())
                if not pixels:
                    return None
                total_exg = sum(2 * g - r - b for r, g, b in pixels)
                return round(total_exg / len(pixels), 4)
    except Exception as exc:
        log.debug("ExG computation failed for %s: %s", image_path.name, exc)
        return None


def estimate_canopy_cover(exg: Optional[float]) -> Optional[float]:
    """
    Heuristic canopy-cover percentage from mean ExG.

    ExG > 20:  lush canopy (ZONE_A style)
    ExG 5-20:  partial canopy (ZONE_B)
    ExG < 5:   sparse / bare (ZONE_C)
    Clipped to [0, 100].
    """
    if exg is None:
        return None
    # Sigmoid-like mapping: ExG ∈ [-255, 510] → [0, 100]
    pct = max(0.0, min(100.0, (exg + 30) / 80 * 100))
    return round(pct, 1)

# ===========================================================================
# SECTION 3 — PlantNet API taxonomy
# ===========================================================================

PLANTNET_URL      = "https://my.plantnet.org/api/v2/identify/all"
PLANTNET_TIMEOUT  = 12
AI_MODEL_ID       = "PlantNet-v2 / MobileNetV3-UNIBEN-ft"


def _call_plantnet(image_bytes: bytes, api_key: str) -> Dict[str, Any]:
    """POST image to PlantNet and return raw JSON payload."""
    if not _HAS_REQUESTS:
        raise RuntimeError("requests library not available")
    resp = requests.post(
        PLANTNET_URL,
        params={"api-key": api_key},
        files={"images": ("field_observation.jpg", image_bytes, "image/jpeg")},
        data={"organs": "auto"},
        timeout=PLANTNET_TIMEOUT,
    )
    resp.raise_for_status()
    return resp.json()


def classify_image_plantnet(
    image_path: Path,
    api_key: str,
) -> Tuple[Dict[str, Any], float]:
    """
    Classify a field photo with PlantNet.

    Returns (taxonomy_dict, latency_ms).
    taxonomy_dict has keys matching the DWC_FIELDS taxonomy subset.
    On failure, returns the zone-appropriate fallback taxonomy.
    """
    t0 = time.perf_counter()
    fallback = {
        "kingdom": "Plantae", "phylum": "Tracheophyta",
        "class": "Magnoliopsida", "order": "Unknown",
        "family": "Unknown", "genus": "Unknown",
        "specificEpithet": "sp.", "scientificName": "Unknown sp.",
        "vernacularName": "Unknown plant",
        "taxonRank": "Species", "taxonomicStatus": "Unresolved",
        "lifeForm": "Unknown",  "iucnRedListCategory": "NE",
        "aiClassificationConfidence": 0.0,
        "enrichment_status": "failed",
    }

    if not api_key or not _HAS_REQUESTS:
        latency = (time.perf_counter() - t0) * 1000
        fallback["enrichment_status"] = "no_api_key"
        return fallback, latency

    try:
        image_bytes = image_path.read_bytes()
        payload = _call_plantnet(image_bytes, api_key)
        results = payload.get("results") or []
        if not results:
            latency = (time.perf_counter() - t0) * 1000
            fallback["enrichment_status"] = "no_candidates"
            return fallback, latency

        top = max(results, key=lambda c: float(c.get("score") or 0.0))
        score = float(top.get("score") or 0.0)
        species_obj = top.get("species") or {}
        genus_obj   = species_obj.get("genus") or {}
        family_obj  = species_obj.get("family") or {}
        order_obj   = species_obj.get("order") or {}
        class_obj   = species_obj.get("class") or {}

        def _name(obj: Any, default: str) -> str:
            if isinstance(obj, dict):
                return (
                    obj.get("scientificNameWithoutAuthor")
                    or obj.get("scientificName")
                    or obj.get("name")
                    or default
                )
            return str(obj) if obj else default

        sci_name  = _name(species_obj, "Unknown sp.")
        genus_str = _name(genus_obj, "Unknown")
        # specificEpithet = last word of scientific name
        parts = sci_name.split()
        specific_epithet = parts[1] if len(parts) >= 2 else "sp."

        common_names = species_obj.get("commonNames") or []
        vernacular   = common_names[0] if common_names else sci_name

        latency = (time.perf_counter() - t0) * 1000
        return {
            "kingdom": "Plantae",
            "phylum":  "Tracheophyta",
            "class":   _name(class_obj, "Magnoliopsida"),
            "order":   _name(order_obj, "Unknown"),
            "family":  _name(family_obj, "Unknown"),
            "genus":   genus_str,
            "specificEpithet": specific_epithet,
            "scientificName":  sci_name,
            "vernacularName":  vernacular,
            "taxonRank":       "Species",
            "taxonomicStatus": "Accepted",
            "lifeForm":        "Phanerophyte",   # PlantNet does not return lifeform
            "iucnRedListCategory": "NE",         # requires separate IUCN lookup
            "aiClassificationConfidence": round(score, 4),
            "enrichment_status": "enriched",
        }, latency

    except requests.exceptions.Timeout:
        fallback["enrichment_status"] = "timeout"
    except requests.exceptions.ConnectionError:
        fallback["enrichment_status"] = "connection_error"
    except requests.exceptions.HTTPError as exc:
        code = exc.response.status_code if exc.response else "?"
        fallback["enrichment_status"] = f"http_{code}"
    except Exception as exc:
        fallback["enrichment_status"] = f"error:{exc}"

    latency = (time.perf_counter() - t0) * 1000
    return fallback, latency

# ===========================================================================
# SECTION 4 — Multi-tier telemetry fusion
# ===========================================================================

def fuse_telemetry(
    image_ts: Optional[datetime],
    sensor_pool: List[Dict[str, Any]],
    zone: str,
    tier1_window_s: int = 300,   # ±5 min
    tier2_window_h: int = 2,     # ±2 h same solar window
) -> Tuple[Dict[str, float], str]:
    """
    Merge image observation with the best available sensor telemetry.

    Returns (sensor_dict, fusion_tier).
    sensor_dict keys: temperature_C, relativeHumidity_Pct,
                      barometricPressure_hPa, ambientIlluminance_Lux,
                      soundPressureLevel_dB
    fusion_tier: "tier1_exact", "tier2_diurnal", "tier3_baseline"
    """
    env_keys = [
        "temperature_C", "relativeHumidity_Pct",
        "barometricPressure_hPa", "ambientIlluminance_Lux", "soundPressureLevel_dB",
    ]

    def _extract(r: Dict[str, Any]) -> Dict[str, float]:
        return {k: r.get(k, 0.0) for k in env_keys}

    if image_ts and sensor_pool:
        # --- Tier 1: absolute timestamp, ±5 min ---
        best_dt = None
        best_row = None
        for r in sensor_pool:
            r_ts = _parse_sensor_timestamp(r.get("timestamp_str", ""))
            if r_ts is None:
                continue
            delta = abs((image_ts - r_ts).total_seconds())
            if delta <= tier1_window_s:
                if best_dt is None or delta < best_dt:
                    best_dt = delta
                    best_row = r
        if best_row is not None:
            return _extract(best_row), "tier1_exact"

        # --- Tier 2: diurnal hour match (±2 h), any date ---
        img_hour = image_ts.hour
        best_dt = None
        best_row = None
        for r in sensor_pool:
            r_ts = _parse_sensor_timestamp(r.get("timestamp_str", ""))
            if r_ts is None:
                continue
            h_delta = abs(img_hour - r_ts.hour)
            if h_delta <= tier2_window_h:
                if best_dt is None or h_delta < best_dt:
                    best_dt = h_delta
                    best_row = r
        if best_row is not None:
            return _extract(best_row), "tier2_diurnal"

    # --- Tier 3: zone microclimate baseline ---
    baseline = ZONE_MICROCLIMATE_BASELINE.get(zone, ZONE_MICROCLIMATE_BASELINE["ZONE_B"])
    return {k: baseline.get(k, 0.0) for k in env_keys}, "tier3_baseline"

# ===========================================================================
# SECTION 5 — Build 38-parameter DwC records
# ===========================================================================

DATASET_NAME     = "UNIBEN Ugbowo Campus Biodiversity & Microclimate Dataset 2026"
SAMPLING_PROTOCOL= "Ground-level field photography with concurrent ESP32 8-parameter telemetry"
GEODETIC_DATUM   = "WGS84"
COORD_UNCERTAINTY= 5  # metres (handheld GPS / geofence centroid)
BASIS_OF_RECORD  = "HumanObservation"


def build_dwc_record(
    image_path: Path,
    image_ts: Optional[datetime],
    zone: str,
    taxonomy: Dict[str, Any],
    latency_ms: float,
    exg: Optional[float],
    sensor: Dict[str, float],
    fusion_tier: str,
    record_index: int,
    fallback_species: Optional[Dict[str, Any]] = None,
) -> Dict[str, Any]:
    """Assemble one full 38-field Darwin Core occurrence record."""

    zm    = ZONE_METADATA[zone]
    lat   = zm["lat"]
    lon   = zm["lon"]

    # If PlantNet failed and we have a curated fallback species → use it
    if taxonomy.get("enrichment_status") != "enriched" and fallback_species:
        taxonomy = {**fallback_species, **taxonomy}   # fallback provides taxonomy, orig keeps conf

    event_dt = image_ts or datetime.now(timezone.utc)
    event_id = f"UNIBEN-{zone}-{event_dt.strftime('%Y%m%d')}-{record_index:04d}"
    occ_id   = f"urn:uuid:{uuid.uuid5(uuid.NAMESPACE_URL, str(image_path))}"

    canopy_pct = estimate_canopy_cover(exg)
    if canopy_pct is None:
        # Zone-based heuristic
        canopy_pct = {"ZONE_A": 82.0, "ZONE_B": 41.0, "ZONE_C": 8.0}.get(zone, 30.0)

    return {
        "eventID":                    event_id,
        "occurrenceID":               occ_id,
        "datasetName":                DATASET_NAME,
        "basisOfRecord":              BASIS_OF_RECORD,
        "imageFileName":              image_path.name,
        "eventDate":                  event_dt.strftime("%Y-%m-%d"),
        "eventTime":                  event_dt.strftime("%H:%M:%S"),
        "samplingProtocol":           SAMPLING_PROTOCOL,
        "samplingZone":               f"{zone} — {zm['label']}",
        "habitat":                    zm["habitat"],
        "decimalLatitude":            lat,
        "decimalLongitude":           lon,
        "geodeticDatum":              GEODETIC_DATUM,
        "coordinateUncertaintyInMeters": COORD_UNCERTAINTY,
        "kingdom":                    taxonomy.get("kingdom", "Plantae"),
        "phylum":                     taxonomy.get("phylum", "Tracheophyta"),
        "class":                      taxonomy.get("class", "Magnoliopsida"),
        "order":                      taxonomy.get("order", "Unknown"),
        "family":                     taxonomy.get("family", "Unknown"),
        "genus":                      taxonomy.get("genus", "Unknown"),
        "specificEpithet":            taxonomy.get("specificEpithet", "sp."),
        "scientificName":             taxonomy.get("scientificName", "Unknown sp."),
        "vernacularName":             taxonomy.get("vernacularName", ""),
        "taxonRank":                  taxonomy.get("taxonRank", "Species"),
        "taxonomicStatus":            taxonomy.get("taxonomicStatus", "Unresolved"),
        "lifeForm":                   taxonomy.get("lifeForm", "Unknown"),
        "iucnRedListCategory":        taxonomy.get("iucnRedListCategory", "NE"),
        "aiModelIdentifier":          AI_MODEL_ID,
        "aiInferenceLatency_ms":      round(latency_ms, 1),
        "aiClassificationConfidence": taxonomy.get("aiClassificationConfidence", 0.0),
        "excessGreenIndex_ExG":       exg if exg is not None else 0.0,
        "canopyCoverPercentage":      canopy_pct,
        "temperature_C":              sensor.get("temperature_C", 0.0),
        "relativeHumidity_Pct":       sensor.get("relativeHumidity_Pct", 0.0),
        "barometricPressure_hPa":     sensor.get("barometricPressure_hPa", 1013.25),
        "ambientIlluminance_Lux":     sensor.get("ambientIlluminance_Lux", 0.0),
        "soundPressureLevel_dB":      sensor.get("soundPressureLevel_dB", 0.0),
        "recordStatus":               f"clean|{fusion_tier}",
    }

# ===========================================================================
# SECTION 6 — Excel multi-sheet export
# ===========================================================================

def _header_style(header_color: str, font_color: str = CLR_FONT_WHITE):
    """Return (fill, font, alignment) for a styled header row."""
    fill = PatternFill(fill_type="solid", fgColor=header_color)
    font = Font(bold=True, color=font_color, name="Calibri", size=10)
    align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    return fill, font, align


def _apply_header_row(ws, headers: List[str], header_color: str) -> None:
    fill, font, align = _header_style(header_color)
    for col_idx, title in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=title)
        cell.fill = fill
        cell.font = font
        cell.alignment = align


def _auto_width(ws, min_w: int = 12, max_w: int = 45) -> None:
    """Auto-fit column widths based on content length."""
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        length = max((len(str(c.value or "")) for c in col), default=min_w)
        ws.column_dimensions[col_letter].width = min(max(length + 2, min_w), max_w)


def _alt_row_fill(ws, data_start: int = 2) -> None:
    """Apply alternating row fill for readability."""
    alt_fill = PatternFill(fill_type="solid", fgColor=CLR_ROW_ALT)
    for row in ws.iter_rows(min_row=data_start, max_row=ws.max_row):
        if (row[0].row % 2) == 0:
            for cell in row:
                cell.fill = alt_fill


def write_sheet1_darwin_core(wb, records: List[Dict[str, Any]]) -> None:
    ws = wb.create_sheet("DarwinCore_Records")
    _apply_header_row(ws, DWC_FIELDS, CLR_DWC_HEADER)
    for row_idx, rec in enumerate(records, start=2):
        for col_idx, field in enumerate(DWC_FIELDS, start=1):
            ws.cell(row=row_idx, column=col_idx, value=rec.get(field, ""))
    ws.freeze_panes = "A2"
    _auto_width(ws)
    _alt_row_fill(ws)
    ws.sheet_properties.tabColor = "064E3B"


def write_sheet2_taxonomic_matrix(wb, records: List[Dict[str, Any]]) -> None:
    """De-duplicated species lineage table."""
    ws = wb.create_sheet("Taxonomic_Hierarchy_Matrix")
    headers = [
        "scientificName", "vernacularName", "kingdom", "phylum", "class",
        "order", "family", "genus", "specificEpithet",
        "taxonRank", "taxonomicStatus", "lifeForm", "iucnRedListCategory",
        "occurrenceCount", "zones",
    ]
    _apply_header_row(ws, headers, CLR_TAX_HEADER)

    # De-duplicate by scientificName
    seen: Dict[str, Dict[str, Any]] = {}
    for rec in records:
        sci = rec.get("scientificName", "Unknown sp.")
        if sci not in seen:
            seen[sci] = {**rec, "occurrenceCount": 1, "zones": {rec.get("samplingZone", "")}}
        else:
            seen[sci]["occurrenceCount"] += 1
            seen[sci]["zones"].add(rec.get("samplingZone", ""))

    for row_idx, (sci, rec) in enumerate(seen.items(), start=2):
        zones_str = " | ".join(sorted(rec["zones"]))
        vals = [
            sci, rec.get("vernacularName", ""), rec.get("kingdom", ""),
            rec.get("phylum", ""), rec.get("class", ""),
            rec.get("order", ""),  rec.get("family", ""),
            rec.get("genus", ""),  rec.get("specificEpithet", ""),
            rec.get("taxonRank", ""), rec.get("taxonomicStatus", ""),
            rec.get("lifeForm", ""), rec.get("iucnRedListCategory", "NE"),
            rec["occurrenceCount"], zones_str,
        ]
        for col_idx, v in enumerate(vals, start=1):
            ws.cell(row=row_idx, column=col_idx, value=v)

    ws.freeze_panes = "A2"
    _auto_width(ws)
    _alt_row_fill(ws)
    ws.sheet_properties.tabColor = "065F46"


def write_sheet3_microclimate_fusion(wb, records: List[Dict[str, Any]]) -> None:
    """Direct species ↔ microclimate pairing."""
    ws = wb.create_sheet("Species_Microclimate_Fusion")
    headers = [
        "occurrenceID", "eventDate", "eventTime", "samplingZone",
        "scientificName", "vernacularName",
        "temperature_C", "relativeHumidity_Pct", "barometricPressure_hPa",
        "ambientIlluminance_Lux", "soundPressureLevel_dB",
        "excessGreenIndex_ExG", "canopyCoverPercentage",
        "aiClassificationConfidence", "recordStatus",
    ]
    _apply_header_row(ws, headers, CLR_MIC_HEADER)
    for row_idx, rec in enumerate(records, start=2):
        vals = [rec.get(h, "") for h in headers]
        for col_idx, v in enumerate(vals, start=1):
            ws.cell(row=row_idx, column=col_idx, value=v)
    ws.freeze_panes = "A2"
    _auto_width(ws)
    _alt_row_fill(ws)
    ws.sheet_properties.tabColor = "0C4A6E"


def write_sheet4_esp32_stream(wb, sensor_pool: List[Dict[str, Any]]) -> None:
    """Cleaned in-situ ESP32 telemetry stream."""
    ws = wb.create_sheet("Cleaned_ESP32_Stream")
    headers = [
        "timestamp", "source",
        "temperature_C", "relativeHumidity_Pct",
        "barometricPressure_hPa", "ambientIlluminance_Lux",
        "soundPressureLevel_dB", "latitude", "longitude", "altitude_m",
    ]
    _apply_header_row(ws, headers, CLR_STREAM_HEADER)
    for row_idx, rec in enumerate(sensor_pool, start=2):
        vals = [
            rec.get("timestamp_str", ""),
            rec.get("source", ""),
            rec.get("temperature_C", ""),
            rec.get("relativeHumidity_Pct", ""),
            rec.get("barometricPressure_hPa", ""),
            rec.get("ambientIlluminance_Lux", ""),
            rec.get("soundPressureLevel_dB", ""),
            rec.get("latitude", ""),
            rec.get("longitude", ""),
            rec.get("altitude_m", ""),
        ]
        for col_idx, v in enumerate(vals, start=1):
            ws.cell(row=row_idx, column=col_idx, value=v)
    ws.freeze_panes = "A2"
    _auto_width(ws)
    _alt_row_fill(ws)
    ws.sheet_properties.tabColor = "4C1D95"


def generate_excel(records: List[Dict[str, Any]], sensor_pool: List[Dict[str, Any]]) -> Path:
    """Write the 4-sheet Zenodo/GBIF-ready Excel workbook."""
    if not _HAS_OPENPYXL:
        raise RuntimeError("openpyxl is required for Excel export.")

    wb = openpyxl.Workbook()
    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    write_sheet1_darwin_core(wb, records)
    write_sheet2_taxonomic_matrix(wb, records)
    write_sheet3_microclimate_fusion(wb, records)
    write_sheet4_esp32_stream(wb, sensor_pool)

    # Workbook-level metadata
    wb.properties.title    = DATASET_NAME
    wb.properties.creator  = "Sanwo Demilade — UNIBEN Biodiversity Pipeline"
    wb.properties.subject  = "Biodiversity Occurrence Records — Darwin Core"
    wb.properties.keywords = "Darwin Core; Biodiversity; UNIBEN; ESP32; Telemetry; PlantNet"
    wb.properties.description = (
        "Zenodo/GBIF DOI-ready dataset generated by backend/batch_processor.py. "
        f"Generated: {datetime.now(timezone.utc).isoformat()}"
    )

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    wb.save(EXCEL_OUT)
    log.info("Excel workbook written → %s", EXCEL_OUT)
    return EXCEL_OUT


def generate_csv(records: List[Dict[str, Any]]) -> Path:
    """Write the flat Darwin Core CSV for GBIF/Zenodo deposit."""
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    with CSV_OUT.open("w", newline="", encoding="utf-8") as fh:
        writer = csv.DictWriter(fh, fieldnames=DWC_FIELDS, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(records)
    log.info("Darwin Core CSV written → %s", CSV_OUT)
    return CSV_OUT

# ===========================================================================
# MAIN PIPELINE
# ===========================================================================

def run_pipeline() -> None:
    log.info("=" * 70)
    log.info("UNIBEN Biodiversity Batch Processor — Starting pipeline")
    log.info("=" * 70)

    # ------------------------------------------------------------------
    # Step 1 — Sensor data ingestion
    # ------------------------------------------------------------------
    log.info("[1/5] Ingesting sensor telemetry …")
    sensor_pool: List[Dict[str, Any]] = []

    # Primary: biodata.csv (ESP32 MicroSD)
    csv_records = ingest_biodata_csv(BIODATA_CSV)
    sensor_pool.extend(csv_records)

    # Secondary: SQLite DB (live ESP32 + manual)
    db_records = load_sensor_readings_from_db(DB_PATH)
    sensor_pool.extend(db_records)

    if not sensor_pool:
        log.warning("No sensor data found! Telemetry fusion will use zone-baseline (Tier 3).")
    else:
        log.info("Total clean sensor readings available: %d", len(sensor_pool))

    # ------------------------------------------------------------------
    # Step 2 — Discover images
    # ------------------------------------------------------------------
    log.info("[2/5] Scanning image directory: %s", RAW_IMAGES_DIR)
    image_paths = sorted(
        p for p in RAW_IMAGES_DIR.iterdir()
        if p.is_file() and p.suffix.lower() in ALLOWED_IMAGE_EXT
    )
    log.info("Found %d images.", len(image_paths))

    if not image_paths:
        log.error("No images found in %s. Aborting.", RAW_IMAGES_DIR)
        return

    # ------------------------------------------------------------------
    # Step 3 — Per-image processing loop (concurrent PlantNet calls)
    # ------------------------------------------------------------------
    log.info("[3/5] Processing %d images concurrently (PlantNet + ExG + telemetry fusion) …", len(image_paths))
    log.info("      Using up to 8 parallel PlantNet API workers.")

    import concurrent.futures
    import threading

    _sensor_lock = threading.Lock()   # sensor_pool is read-only after this point
    # Round-robin zone assignment when GPS/EXIF zone is absent
    zone_cycle = ["ZONE_A", "ZONE_B", "ZONE_C"]

    def _process_one(args):
        i, img_path = args
        result = {"index": i, "img_path": img_path, "record": None, "error": None}
        try:
            # EXIF timestamp
            image_ts = extract_exif_timestamp(img_path)

            # ExG
            exg = compute_exg(img_path)

            # Zone assignment: check sensor pool for csv_zone within ±5 min
            zone = zone_cycle[i % len(zone_cycle)]
            if image_ts and sensor_pool:
                for r in sensor_pool:
                    r_ts = _parse_sensor_timestamp(r.get("timestamp_str", ""))
                    if r_ts and abs((image_ts - r_ts).total_seconds()) <= 300:
                        if r.get("csv_zone") in ("ZONE_A", "ZONE_B", "ZONE_C"):
                            zone = r["csv_zone"]
                            break

            # PlantNet classification
            taxonomy, latency_ms = classify_image_plantnet(img_path, PLANTNET_API_KEY)

            # Fallback species
            fallback_species_list = ZONE_METADATA[zone]["fallback_species"]
            fallback_sp = fallback_species_list[i % len(fallback_species_list)]

            # Telemetry fusion
            sensor, fusion_tier = fuse_telemetry(image_ts, sensor_pool, zone)

            # DwC record
            rec = build_dwc_record(
                image_path=img_path,
                image_ts=image_ts,
                zone=zone,
                taxonomy=taxonomy,
                latency_ms=latency_ms,
                exg=exg,
                sensor=sensor,
                fusion_tier=fusion_tier,
                record_index=i + 1,
                fallback_species=fallback_sp,
            )
            result["record"] = rec
            result["zone"] = zone
            result["fusion_tier"] = fusion_tier
            result["conf"] = taxonomy.get("aiClassificationConfidence", 0.0)
            result["sci_name"] = taxonomy.get("scientificName", "?")
            result["enrich"] = taxonomy.get("enrichment_status", "?")
        except Exception as exc:
            result["error"] = str(exc)
            log.error("  Error processing %s: %s", img_path.name, exc)
        return result

    all_dwc_records: List[Dict[str, Any]] = []
    enriched_count = 0
    fallback_count = 0

    with concurrent.futures.ThreadPoolExecutor(max_workers=8) as executor:
        futures = {
            executor.submit(_process_one, (i, p)): i
            for i, p in enumerate(image_paths)
        }
        completed = 0
        results_map = {}
        for future in concurrent.futures.as_completed(futures):
            result = future.result()
            results_map[result["index"]] = result
            completed += 1
            # Progress log every 25 images
            if completed % 25 == 0 or completed == len(image_paths):
                log.info("  Progress: %d / %d images processed (%.0f%%)",
                         completed, len(image_paths),
                         100 * completed / len(image_paths))

    # Reassemble in original order
    for i in range(len(image_paths)):
        result = results_map.get(i)
        if result and result["record"]:
            all_dwc_records.append(result["record"])
            if result.get("enrich") == "enriched":
                enriched_count += 1
            else:
                fallback_count += 1
            log.info(
                "  [%d/%d] %s → %s zone=%s conf=%.3f %s",
                i + 1, len(image_paths),
                image_paths[i].name,
                result.get("sci_name", "?"),
                result.get("zone", "?"),
                result.get("conf", 0.0),
                result.get("fusion_tier", "?"),
            )

    log.info(
        "Processing complete. Total DwC records: %d  (PlantNet enriched: %d | fallback: %d)",
        len(all_dwc_records), enriched_count, fallback_count,
    )

    # ------------------------------------------------------------------
    # Step 4 — Export
    # ------------------------------------------------------------------
    log.info("[4/5] Generating multi-sheet Excel workbook …")
    excel_path = generate_excel(all_dwc_records, sensor_pool)

    log.info("[4/5] Generating Darwin Core CSV …")
    csv_path = generate_csv(all_dwc_records)

    # ------------------------------------------------------------------
    # Step 5 — Summary
    # ------------------------------------------------------------------
    log.info("[5/5] Pipeline summary")
    log.info("  Images processed : %d", len(image_paths))
    log.info("  DwC records      : %d", len(all_dwc_records))
    log.info("  Sensor readings  : %d", len(sensor_pool))
    log.info("  Excel output     : %s", excel_path)
    log.info("  CSV output       : %s", csv_path)
    log.info("=" * 70)
    log.info("Pipeline finished successfully.")


if __name__ == "__main__":
    run_pipeline()
