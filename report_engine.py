"""
report_engine.py — Excel & Reporting Layer
==========================================
UNIBEN Biodiversity Pipeline · Root Module

Generates structured multi-sheet Excel (.xlsx) workbooks from the live
data/biodiversity.db database and archives them in data/exports/.

Sheets produced
---------------
  1. Master Linked Encounters   — timestamp-aligned drone + ground + taxonomy
  2. Continuous Telemetry Logs  — full 5-parameter sensor time-series
  3. Taxonomic Classification Inventory — species counts, confidence, group

Usage
-----
    # CLI one-shot:
    python report_engine.py [session_id]

    # Import into FastAPI:
    from report_engine import ReportEngine
    engine = ReportEngine()
    path = engine.generate_excel_spreadsheet(session_id="all")
"""

from __future__ import annotations

import logging
import sqlite3
from datetime import datetime, timezone
from pathlib import Path
from typing import Optional

log = logging.getLogger("report_engine")
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
)

# ---------------------------------------------------------------------------
# Dependency guards
# ---------------------------------------------------------------------------

try:
    import pandas as pd
    _HAS_PANDAS = True
except ImportError:
    _HAS_PANDAS = False

try:
    import openpyxl                           # noqa: F401  (checked at runtime)
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side,
    )
    from openpyxl.utils import get_column_letter
    _HAS_OPENPYXL = True
except ImportError:
    _HAS_OPENPYXL = False

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

BASE_DIR    = Path(__file__).resolve().parent
DB_PATH     = BASE_DIR / "data" / "biodiversity.db"
EXPORTS_DIR = BASE_DIR / "data" / "exports"

# Colour palette (ARGB hex without #)
CLR_HEADER_DARK   = "FF064E3B"   # forest green
CLR_HEADER_MID    = "FF065F46"
CLR_HEADER_LIGHT  = "FF34D399"
CLR_ROW_ALT       = "FFE8FDF5"
CLR_ROW_PLAIN     = "FFFFFFFF"
CLR_FONT_LIGHT     = "FFFFFFFF"
CLR_FONT_DARK      = "FF111827"

SENSOR_FIELDS = [
    "temperature_c", "humidity_percent", "pressure_hPa",
    "light_lux", "sound_db",
]
SENSOR_LABELS = {
    "temperature_c":    "Temp (°C)",
    "humidity_percent": "Humidity (%)",
    "pressure_hPa":     "Pressure (hPa)",
    "light_lux":        "Light (Lux)",
    "sound_db":         "Sound (dB)",
}


# ===========================================================================
# Styling helpers
# ===========================================================================

def _header_font(bold: bool = True) -> "Font":
    return Font(name="Calibri", bold=bold, color=CLR_FONT_LIGHT, size=10)


def _body_font() -> "Font":
    return Font(name="Calibri", size=10, color=CLR_FONT_DARK)


def _header_fill(hex_color: str = CLR_HEADER_DARK) -> "PatternFill":
    return PatternFill(fill_type="solid", fgColor=hex_color)


def _alt_fill(row_idx: int) -> "PatternFill":
    color = CLR_ROW_ALT if row_idx % 2 == 0 else CLR_ROW_PLAIN
    return PatternFill(fill_type="solid", fgColor=color)


def _thin_border() -> "Border":
    thin = Side(style="thin", color="FFD1FAE5")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def _centre() -> "Alignment":
    return Alignment(horizontal="center", vertical="center", wrap_text=False)


def _left() -> "Alignment":
    return Alignment(horizontal="left", vertical="center", wrap_text=False)


def _apply_header_row(ws, headers: list[str], fill_color: str = CLR_HEADER_DARK) -> None:
    """Write styled header cells to row 1 of a worksheet."""
    for col_idx, header in enumerate(headers, start=1):
        cell           = ws.cell(row=1, column=col_idx, value=header)
        cell.font      = _header_font()
        cell.fill      = _header_fill(fill_color)
        cell.alignment = _centre()
        cell.border    = _thin_border()


def _apply_data_row(ws, row_idx: int, values: list, right_align_cols: set = None) -> None:
    """Write styled data cells to row row_idx."""
    right_align_cols = right_align_cols or set()
    for col_idx, value in enumerate(values, start=1):
        cell           = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.font      = _body_font()
        cell.fill      = _alt_fill(row_idx)
        cell.alignment = _centre() if col_idx in right_align_cols else _left()
        cell.border    = _thin_border()


def _autofit_columns(ws, min_width: int = 10, max_width: int = 55) -> None:
    """Auto-size column widths based on max content length."""
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        max_len    = max(
            (len(str(cell.value)) if cell.value is not None else 0)
            for cell in col
        )
        ws.column_dimensions[col_letter].width = max(min_width, min(max_width, max_len + 4))


def _freeze_header(ws) -> None:
    ws.freeze_panes = ws.cell(row=2, column=1)


# ===========================================================================
# Report Engine
# ===========================================================================

class ReportEngine:
    """
    Generates multi-sheet Excel workbooks from the live biodiversity database.

    Parameters
    ----------
    db_path    : Path to data/biodiversity.db
    exports_dir: Output directory for generated .xlsx files
    """

    def __init__(
        self,
        db_path:     Path | str = DB_PATH,
        exports_dir: Path | str = EXPORTS_DIR,
    ) -> None:
        self.db_path     = Path(db_path)
        self.exports_dir = Path(exports_dir)
        self.exports_dir.mkdir(parents=True, exist_ok=True)

    # ------------------------------------------------------------------
    # Public API
    # ------------------------------------------------------------------

    def generate_excel_spreadsheet(self, session_id: str = "all") -> Path:
        """
        Build and save a multi-sheet .xlsx workbook for the given session_id.

        Parameters
        ----------
        session_id : "all" to include all records, or a specific sync_session_id
                     string to filter to a single field run.

        Returns
        -------
        Path to the saved .xlsx file.

        Raises
        ------
        ImportError   — if pandas or openpyxl are not installed.
        RuntimeError  — if the database cannot be read.
        """
        if not _HAS_PANDAS:
            raise ImportError("pandas is required to generate reports.")
        if not _HAS_OPENPYXL:
            raise ImportError(
                "openpyxl is required to generate Excel reports. "
                "Install it with: pip install openpyxl"
            )

        log.info("ReportEngine: building Excel workbook (session_id=%s)", session_id)

        sensor_df, image_df, linked_df = self._load_data(session_id)
        taxonomy_df                    = self._build_taxonomy_inventory(image_df)

        import openpyxl
        wb = openpyxl.Workbook()

        # Remove default sheet
        wb.remove(wb.active)

        self._sheet_master_encounters(wb, linked_df)
        self._sheet_telemetry_logs(wb, sensor_df)
        self._sheet_taxonomy_inventory(wb, taxonomy_df)

        timestamp = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")
        safe_sid  = session_id.replace(" ", "_").replace("/", "-")[:30]
        filename  = f"biodiversity_report_{safe_sid}_{timestamp}.xlsx"
        out_path  = self.exports_dir / filename

        wb.save(str(out_path))
        log.info("ReportEngine: saved → %s  (%d KB)", out_path, out_path.stat().st_size // 1024)
        return out_path

    # ------------------------------------------------------------------
    # Data loading
    # ------------------------------------------------------------------

    def _load_data(self, session_id: str):
        """
        Pull sensor_readings, drone_patches, and field_observations
        from SQLite. Merges them for the linked sheet.
        """
        if not self.db_path.exists():
            empty = pd.DataFrame()
            return empty, empty, empty

        conn = sqlite3.connect(self.db_path)
        conn.row_factory = sqlite3.Row

        try:
            # Sensor readings
            sensor_sql  = "SELECT * FROM sensor_readings ORDER BY observed_at ASC"
            sensor_df   = pd.read_sql_query(sensor_sql, conn)

            # Drone patches
            drone_sql   = "SELECT * FROM drone_patches ORDER BY flight_timestamp ASC"
            drone_df    = pd.read_sql_query(drone_sql, conn)

            # Field observations
            obs_sql     = "SELECT * FROM field_observations ORDER BY date ASC, time ASC"
            obs_df      = pd.read_sql_query(obs_sql, conn)

        finally:
            conn.close()

        # Merge for linked encounters sheet (drone_patches + field_observations)
        if not obs_df.empty and not drone_df.empty:
            linked_df = pd.merge(
                obs_df,
                drone_df,
                on="drone_id",
                how="left",
                suffixes=("_obs", "_drone"),
            )
        elif not obs_df.empty:
            linked_df = obs_df.copy()
        else:
            linked_df = pd.DataFrame()

        return sensor_df, drone_df, linked_df

    # ------------------------------------------------------------------
    # Taxonomy inventory builder
    # ------------------------------------------------------------------

    def _build_taxonomy_inventory(self, linked_df: "pd.DataFrame") -> "pd.DataFrame":
        """
        Aggregate field_observations into species count statistics.
        """
        if linked_df.empty or "common_name" not in linked_df.columns:
            return pd.DataFrame(columns=[
                "Species / Label", "Encounter Count",
                "Mean Confidence (%)", "Min Confidence (%)", "Max Confidence (%)",
            ])

        df = linked_df.copy()
        df["common_name"] = df["common_name"].fillna("Unclassified")

        grp = df.groupby("common_name").agg(
            count          = ("common_name", "count"),
            mean_conf      = ("annotation_confidence", "mean"),
            min_conf       = ("annotation_confidence", "min"),
            max_conf       = ("annotation_confidence", "max"),
        ).reset_index()

        grp["mean_conf"] = (grp["mean_conf"] * 100).round(2)
        grp["min_conf"]  = (grp["min_conf"]  * 100).round(2)
        grp["max_conf"]  = (grp["max_conf"]  * 100).round(2)

        grp = grp.sort_values("count", ascending=False).reset_index(drop=True)
        grp.columns = [
            "Species / Label", "Encounter Count",
            "Mean Confidence (%)", "Min Confidence (%)", "Max Confidence (%)",
        ]
        return grp

    # ------------------------------------------------------------------
    # Sheet builders
    # ------------------------------------------------------------------

    def _sheet_master_encounters(self, wb, df: "pd.DataFrame") -> None:
        """
        Sheet 1: Master Linked Encounters
        Side-by-side rows: drone patches and 38-parameter field observations.
        """
        ws        = wb.create_sheet("Master Linked Encounters")
        ws.sheet_properties.tabColor = "064E3B"

        headers = [
            "Row #", "Observation ID", "Drone ID",
            "Drone Image Path", "Campus Zone", "Flight Timestamp",
            "Ground Image Path", "Category", "Kingdom", "Phylum", "Class", "Order",
            "Family", "Genus", "Species", "Common Name", "Local Name",
            "GPS Lat", "GPS Long", "Habitat Type", "Count", "Abundance Class",
            "Life Stage", "Sex", "Health Status", "Behaviour", "Microhabitat",
            "Temp (°C)", "Humidity (%)", "Light (Lux)", "Pressure (hPa)", "Sound (dB)",
            "Wind Speed (m/s)", "Rainfall (mm)", "IUCN Status", "Origin Status",
            "Annotation Confidence", "ML Subset", "Observer ID", "Date", "Time", "Week No",
        ]
        _apply_header_row(ws, headers, CLR_HEADER_DARK)
        _freeze_header(ws)

        if df.empty:
            ws.cell(row=2, column=1, value="No encounters recorded yet.")
            _autofit_columns(ws)
            return

        for row_idx, (_, row) in enumerate(df.iterrows(), start=2):
            vals = [
                row_idx - 1,
                row.get("obs_id", "—"),
                row.get("drone_id", "—"),
                row.get("drone_image_path", "—"),
                row.get("campus_zone", "—"),
                _ts(row.get("flight_timestamp")),
                row.get("ground_image_path", "—"),
                row.get("category", "—"),
                row.get("kingdom", "—"),
                row.get("phylum", "—"),
                row.get("class", "—"),
                row.get("order_name", "—"),
                row.get("family", "—"),
                row.get("genus", "—"),
                row.get("species", "—"),
                row.get("common_name", "—"),
                row.get("local_name", "—"),
                _fmt(row.get("gps_lat"), 6),
                _fmt(row.get("gps_long"), 6),
                row.get("habitat_type", "—"),
                row.get("count", "—"),
                row.get("abundance_class", "—"),
                row.get("life_stage", "—"),
                row.get("sex", "—"),
                row.get("health_status", "—"),
                row.get("behaviour", "—"),
                row.get("microhabitat", "—"),
                _fmt(row.get("ambient_temp_c")),
                _fmt(row.get("rel_humidity_pct")),
                _fmt(row.get("light_lux")),
                _fmt(row.get("atmospheric_pressure_hpa")),
                _fmt(row.get("ambient_sound_db")),
                _fmt(row.get("wind_speed_ms")),
                _fmt(row.get("rainfall_mm")),
                row.get("iucn_status", "—"),
                row.get("origin_status", "—"),
                f"{float(row['annotation_confidence']) * 100:.1f}%" if row.get("annotation_confidence") is not None else "—",
                row.get("ml_subset", "—"),
                row.get("observer_id", "—"),
                row.get("date", "—"),
                row.get("time", "—"),
                row.get("week_no", "—"),
            ]
            _apply_data_row(ws, row_idx, vals)

        _autofit_columns(ws)
        log.info("Sheet 'Master Linked Encounters': %d data rows.", len(df))

    def _sheet_telemetry_logs(self, wb, df: "pd.DataFrame") -> None:
        """
        Sheet 2: Continuous Telemetry Logs
        Full 5-parameter time-series from sensor_readings.
        """
        ws = wb.create_sheet("Continuous Telemetry Logs")
        ws.sheet_properties.tabColor = "065F46"

        headers = [
            "Row #", "Record ID", "Device ID",
            "Temp (°C)", "Humidity (%)", "Pressure (hPa)",
            "Light (Lux)", "Sound (dB)",
            "Latitude", "Longitude", "Altitude (m)",
            "Observed At", "Received At", "Notes",
        ]
        _apply_header_row(ws, headers, CLR_HEADER_MID)
        _freeze_header(ws)

        if df.empty:
            ws.cell(row=2, column=1, value="No telemetry records found.")
            _autofit_columns(ws)
            return

            vals = [
                row_idx - 1,
                row.get("id"),
                row.get("device_id",        "—"),
                _fmt(row.get("temperature_c")),
                _fmt(row.get("humidity_percent")),
                _fmt(row.get("pressure_hPa")),
                _fmt(row.get("light_lux")),
                _fmt(row.get("sound_db")),
                _fmt(row.get("latitude")),
                _fmt(row.get("longitude")),
                _fmt(row.get("altitude_m")),
                _ts(row.get("observed_at")),
                _ts(row.get("received_at")),
                row.get("notes",            ""),
                row.get("data_source",      "—"),
            ]
            _apply_data_row(ws, row_idx, vals)

        _autofit_columns(ws)
        log.info("Sheet 'Continuous Telemetry Logs': %d data rows.", len(df))

    def _sheet_taxonomy_inventory(self, wb, df: "pd.DataFrame") -> None:
        """
        Sheet 3: Taxonomic Classification Inventory
        Species counts, confidence percentages, group types.
        """
        ws = wb.create_sheet("Taxonomic Classification Inventory")
        ws.sheet_properties.tabColor = "34D399"

        headers = list(df.columns) if not df.empty else [
            "Species / Label", "Encounter Count",
            "Mean Confidence (%)", "Min Confidence (%)", "Max Confidence (%)",
        ]
        _apply_header_row(ws, headers, CLR_HEADER_DARK)
        _freeze_header(ws)

        if df.empty:
            ws.cell(row=2, column=1, value="No classification records found.")
            _autofit_columns(ws)
            return

        for row_idx, (_, row) in enumerate(df.iterrows(), start=2):
            _apply_data_row(ws, row_idx, list(row), right_align_cols={2, 3, 4, 5})

        _autofit_columns(ws)
        log.info("Sheet 'Taxonomic Classification Inventory': %d species.", len(df))

    # ------------------------------------------------------------------
    # Academic PDF Generator
    # ------------------------------------------------------------------
    def generate_academic_pdf(self) -> Path:
        """
        Generate a summary academic PDF report.
        """
        try:
            from reportlab.lib.pagesizes import letter
            from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib import colors
        except ImportError:
            raise ImportError("reportlab is required to generate PDF reports. Install it with: pip install reportlab")
            
        timestamp = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")
        filename  = f"biodiversity_academic_report_{timestamp}.pdf"
        out_path  = self.exports_dir / filename
        
        doc = SimpleDocTemplate(str(out_path), pagesize=letter)
        styles = getSampleStyleSheet()
        elements = []
        
        # Title
        title_style = styles['Title']
        elements.append(Paragraph("UNIBEN Biodiversity Pipeline - Academic ML Insights", title_style))
        elements.append(Spacer(1, 12))
        
        # Summary
        body_style = styles['Normal']
        elements.append(Paragraph(f"Generated at: {timestamp}", body_style))
        elements.append(Spacer(1, 12))
        
        sensor_df, drone_df, linked_df = self._load_data("all")
        elements.append(Paragraph(f"Total Environmental Readings: {len(sensor_df)}", body_style))
        elements.append(Paragraph(f"Total Drone Patches: {len(drone_df)}", body_style))
        elements.append(Paragraph(f"Total Field Observations: {len(linked_df)}", body_style))
        elements.append(Spacer(1, 12))
        
        # Add a simple table of species count
        taxonomy_df = self._build_taxonomy_inventory(linked_df)
        if not taxonomy_df.empty:
            elements.append(Paragraph("Species Inventory Summary", styles['Heading2']))
            data = [taxonomy_df.columns.tolist()] + taxonomy_df.head(10).values.tolist()
            t = Table(data)
            t.setStyle(TableStyle([
                ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#064E3B')),
                ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
                ('ALIGN', (0,0), (-1,-1), 'CENTER'),
                ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
                ('FONTSIZE', (0,0), (-1,0), 10),
                ('BOTTOMPADDING', (0,0), (-1,0), 12),
                ('BACKGROUND', (0,1), (-1,-1), colors.beige),
                ('GRID', (0,0), (-1,-1), 1, colors.black),
            ]))
            elements.append(t)

        doc.build(elements)
        log.info("ReportEngine: PDF saved → %s", out_path)
        return out_path


# ===========================================================================
# Formatting helpers
# ===========================================================================

def _fmt(val, decimals: int = 2) -> str | float | None:
    """Format a float value for display, returning '—' for None."""
    if val is None:
        return "—"
    try:
        return round(float(val), decimals)
    except (TypeError, ValueError):
        return str(val)


def _ts(val) -> str:
    """Format a timestamp string cleanly."""
    if val is None:
        return "—"
    try:
        dt = pd.to_datetime(val, utc=True)
        return dt.strftime("%Y-%m-%d %H:%M:%S UTC")
    except Exception:
        return str(val)


# ===========================================================================
# CLI entry point
# ===========================================================================

if __name__ == "__main__":
    import sys
    sid    = sys.argv[1] if len(sys.argv) > 1 else "all"
    engine = ReportEngine()
    path   = engine.generate_excel_spreadsheet(session_id=sid)
    print(f"Report saved: {path}")
